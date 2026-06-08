from datetime import datetime
from io import BytesIO
from typing import Any, AsyncIterable, Iterable, Mapping, Optional
from urllib.parse import quote

from time_utils import beijing_filename_timestamp

EXPORT_HEADERS = (
    "流水号",
    "申领日期",
    "申领部门",
    "经办人",
    "供应商",
    "物品名称",
    "数量",
    "单价",
    "状态",
    "到货日期",
    "分发日期",
    "签收备注",
)
EXPORT_COLUMN_WIDTHS = (18, 12, 24, 12, 18, 28, 10, 10, 12, 12, 12, 28)
EXPORT_FALLBACK_FILENAME = "procure_lite_export.xlsx"
EXPORT_DISPLAY_NAME_PREFIX = "采购台账"
SUPPLIER_EXPORT_FALLBACK_FILENAME = "supplier_purchase_report.xlsx"
SUPPLIER_EXPORT_DISPLAY_NAME_PREFIX = "供应商采购报表"


class ExportDependencyError(RuntimeError):
    """Export dependency is unavailable."""


def _build_item_row(item: Mapping[str, Any]) -> list[Any]:
    return [
        item.get("serial_number", ""),
        item.get("request_date", ""),
        item.get("department", ""),
        item.get("handler", ""),
        item.get("supplier_name_snapshot", ""),
        item.get("item_name", ""),
        item.get("quantity", ""),
        "" if item.get("unit_price") is None else item.get("unit_price"),
        item.get("status", ""),
        item.get("arrival_date", ""),
        item.get("distribution_date", ""),
        item.get("signoff_note", ""),
    ]


def _append_rows(sheet, headers: list[str], rows: Iterable[Iterable[Any]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(list(row))


def build_items_excel_stream(items: Iterable[Mapping[str, Any]]) -> BytesIO:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        raise ExportDependencyError(
            "缺少 openpyxl 依赖，请先安装 requirements.txt"
        ) from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "采购记录"
    sheet.append(list(EXPORT_HEADERS))

    for item in items:
        sheet.append(_build_item_row(item))

    for idx, width in enumerate(EXPORT_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


async def build_items_excel_stream_async(
    items: AsyncIterable[Mapping[str, Any]],
) -> BytesIO:
    """低内存版：逐行消费异步生成器，省去全量列表中间对象。"""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        raise ExportDependencyError(
            "缺少 openpyxl 依赖，请先安装 requirements.txt"
        ) from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "采购记录"
    sheet.append(list(EXPORT_HEADERS))

    async for item in items:
        sheet.append(_build_item_row(item))

    for idx, width in enumerate(EXPORT_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_supplier_report_excel_stream(
    report: Mapping[str, Any], *, mode: str = "full"
) -> BytesIO:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        raise ExportDependencyError(
            "缺少 openpyxl 依赖，请先安装 requirements.txt"
        ) from exc

    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "供应商汇总"
    summary = report.get("summary") or {}
    summary_sheet.append(["指标", "值"])
    summary_sheet.append(["记录总数", summary.get("total_records", 0)])
    summary_sheet.append(["已归属供应商记录", summary.get("assigned_records", 0)])
    summary_sheet.append(["未归属供应商记录", summary.get("unassigned_records", 0)])
    summary_sheet.append(["供应商数", summary.get("supplier_count", 0)])
    summary_sheet.append(["采购总额", summary.get("total_amount", 0)])
    summary_sheet.append(["已归属金额", summary.get("assigned_amount", 0)])
    summary_sheet.append(["未归属金额", summary.get("unassigned_amount", 0)])
    summary_sheet.append([])
    _append_rows(
        summary_sheet,
        [
            "供应商",
            "供应商ID",
            "记录数",
            "商品数",
            "采购数量",
            "采购总额",
            "最近采购日期",
        ],
        [
            [
                row.get("supplier_name", ""),
                row.get("supplier_id", ""),
                row.get("record_count", 0),
                row.get("item_count", 0),
                row.get("total_quantity", 0),
                row.get("total_amount", 0),
                row.get("latest_request_date", ""),
            ]
            for row in (report.get("top_suppliers") or [])
        ],
    )

    item_sheet = workbook.create_sheet("供应商-商品明细")
    _append_rows(
        item_sheet,
        [
            "供应商",
            "供应商ID",
            "物品名称",
            "记录数",
            "采购数量",
            "采购总额",
            "最近采购日期",
        ],
        [
            [
                row.get("supplier_name", ""),
                row.get("supplier_id", ""),
                row.get("item_name", ""),
                row.get("record_count", 0),
                row.get("total_quantity", 0),
                row.get("total_amount", 0),
                row.get("latest_request_date", ""),
            ]
            for row in (report.get("supplier_items") or [])
        ],
    )

    if mode in {"full", "monthly"}:
        trend_sheet = workbook.create_sheet("月度走势")
        _append_rows(
            trend_sheet,
            ["月份", "供应商", "供应商ID", "记录数", "采购数量", "采购总额"],
            [
                [
                    row.get("month", ""),
                    row.get("supplier_name", ""),
                    row.get("supplier_id", ""),
                    row.get("record_count", 0),
                    row.get("total_quantity", 0),
                    row.get("total_amount", 0),
                ]
                for row in (report.get("monthly_trend") or [])
            ],
        )

    if mode in {"full", "quarterly"}:
        quarterly_sheet = workbook.create_sheet("季度汇总")
        _append_rows(
            quarterly_sheet,
            ["季度", "供应商", "供应商ID", "记录数", "采购数量", "采购总额"],
            [
                [
                    row.get("quarter", ""),
                    row.get("supplier_name", ""),
                    row.get("supplier_id", ""),
                    row.get("record_count", 0),
                    row.get("total_quantity", 0),
                    row.get("total_amount", 0),
                ]
                for row in (report.get("quarterly_trend") or [])
            ],
        )

    if mode in {"full", "yearly"}:
        yearly_sheet = workbook.create_sheet("年度汇总")
        _append_rows(
            yearly_sheet,
            ["年份", "供应商", "供应商ID", "记录数", "商品数", "采购数量", "采购总额"],
            [
                [
                    row.get("year", ""),
                    row.get("supplier_name", ""),
                    row.get("supplier_id", ""),
                    row.get("record_count", 0),
                    row.get("item_count", 0),
                    row.get("total_quantity", 0),
                    row.get("total_amount", 0),
                ]
                for row in (report.get("yearly_summary") or [])
            ],
        )

    unassigned_sheet = workbook.create_sheet("未归属供应商")
    _append_rows(
        unassigned_sheet,
        [
            "记录ID",
            "流水号",
            "申领日期",
            "部门",
            "经办人",
            "物品名称",
            "数量",
            "单价",
            "状态",
        ],
        [
            [
                row.get("id", ""),
                row.get("serial_number", ""),
                row.get("request_date", ""),
                row.get("department", ""),
                row.get("handler", ""),
                row.get("item_name", ""),
                row.get("quantity", 0),
                row.get("unit_price", 0),
                row.get("status", ""),
            ]
            for row in (report.get("unassigned_items") or [])
        ],
    )

    for sheet in workbook.worksheets:
        for idx, column_cells in enumerate(sheet.columns, start=1):
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(idx)].width = min(
                max(12, max_length + 2), 28
            )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_export_content_disposition(
    now: Optional[datetime] = None,
    *,
    fallback_filename: str = EXPORT_FALLBACK_FILENAME,
    display_name_prefix: str = EXPORT_DISPLAY_NAME_PREFIX,
) -> str:
    timestamp = beijing_filename_timestamp(now)
    filename = f"{display_name_prefix}_{timestamp}.xlsx"
    encoded_filename = quote(filename)
    return (
        f'attachment; filename="{fallback_filename}"; '
        f"filename*=UTF-8''{encoded_filename}"
    )
